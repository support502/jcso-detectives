"""
Vercel serverless function: /api/timesheet
Handles pull/save/reset/export actions for the supervisor timesheet feature.
"""
from http.server import BaseHTTPRequestHandler
import json, os, io, datetime, traceback

# ── Supabase client (lazy-init) ─────────────────────────────────────────────
_supabase = None

def get_supabase():
    global _supabase
    if _supabase is None:
        from supabase import create_client
        url = os.environ.get('SUPABASE_URL')
        key = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')
        if not url or not key:
            raise RuntimeError(
                'Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY env vars. '
                'Set them in Vercel project settings.'
            )
        _supabase = create_client(url, key)
    return _supabase

# ── Federal holidays (observed dates) ───────────────────────────────────────
# Add new years by appending to the dict.
FEDERAL_HOLIDAYS = {
    2026: [
        datetime.date(2026, 1, 1),   # New Year's Day
        datetime.date(2026, 1, 19),  # MLK Day
        datetime.date(2026, 2, 16),  # Presidents Day
        datetime.date(2026, 5, 25),  # Memorial Day
        datetime.date(2026, 6, 19),  # Juneteenth
        datetime.date(2026, 7, 3),   # Independence Day (observed, 7/4 is Sat)
        datetime.date(2026, 9, 7),   # Labor Day
        datetime.date(2026, 10, 12), # Columbus Day
        datetime.date(2026, 11, 11), # Veterans Day
        datetime.date(2026, 11, 26), # Thanksgiving
        datetime.date(2026, 12, 25), # Christmas
    ],
}

def _all_holidays():
    """Flat set of all holiday dates across all years."""
    s = set()
    for dates in FEDERAL_HOLIDAYS.values():
        s.update(dates)
    return s

# ── Template path ───────────────────────────────────────────────────────────
TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), '..', 'templates')
MASTER_BLANK = os.path.join(TEMPLATE_DIR, 'MASTER_BLANK.xlsx')

# ── Template layout helpers ────────────────────────────────────────────────

# 14 day columns: C(Sun wk1) .. P(Sat wk2)  →  column indices 3..16
DAY_COL_START = 3   # C
DAY_COL_END   = 16  # P


def _find_blocks(ws):
    """Scan a worksheet for detective blocks.

    Returns a list of (rg_row, code_rows, total_row) tuples found by
    locating 'RG' in column B (code column) and 'TOTAL' in column R.
    """
    rg_rows = []
    total_rows = []
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 2).value == 'RG':       # col B
            rg_rows.append(row)
        if ws.cell(row, 18).value == 'TOTAL':    # col R
            total_rows.append(row)

    blocks = []
    for rg, total in zip(rg_rows, total_rows):
        code_rows = list(range(rg + 1, total))
        blocks.append((rg, code_rows, total))
    return blocks


def _fill_detective_block(ws, rg_row, code_rows_list, ts_rows, detective_name):
    """Fill one detective block on a worksheet.

    Writes only cell values in columns A-P.  Columns Q and R contain
    SUM / =B formulas in the template and are left untouched so Excel
    recalculates them on open.  The TOTAL row is also formula-driven.
    """
    from openpyxl.styles import Font

    # Name in col A of RG row — full name, bold
    name_cell = ws.cell(rg_row, 1)
    name_cell.value = detective_name
    existing = name_cell.font
    name_cell.font = Font(
        name=existing.name,
        size=existing.size,
        bold=True,
    )

    # Fill RG hours (cols C-P)
    for i, tr in enumerate(ts_rows):
        val = float(tr.get('reg_hours', 0))
        ws.cell(rg_row, DAY_COL_START + i).value = val if val else None

    # Gather unique codes in order of appearance
    code_order = []
    for tr in ts_rows:
        for cr in (tr.get('code_rows') or []):
            code = cr.get('code', '')
            if code and code not in code_order:
                code_order.append(code)

    # Build per-code day arrays
    code_days = {c: [0.0] * 14 for c in code_order}
    for i, tr in enumerate(ts_rows):
        for cr in (tr.get('code_rows') or []):
            code = cr.get('code', '')
            hrs = float(cr.get('hours', 0))
            if code in code_days:
                code_days[code][i] += hrs

    # Write code rows (code name in col B, hours in cols C-P)
    for code_idx, code in enumerate(code_order):
        if code_idx >= len(code_rows_list):
            break
        row_num = code_rows_list[code_idx]
        ws.cell(row_num, 2).value = code  # B
        for day_i in range(14):
            val = code_days[code][day_i]
            ws.cell(row_num, DAY_COL_START + day_i).value = val if val else None


def _clear_block_defaults(ws, rg_row):
    """Clear the default placeholder 8s from an unused RG row (cols C-P)."""
    for col in range(DAY_COL_START, DAY_COL_END + 1):
        ws.cell(rg_row, col).value = None


def _cors_headers(h):
    h.send_header('Access-Control-Allow-Origin', '*')
    h.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
    h.send_header('Access-Control-Allow-Headers', 'Content-Type')


def _json_response(h, status, data):
    h.send_response(status)
    h.send_header('Content-Type', 'application/json')
    _cors_headers(h)
    h.end_headers()
    h.wfile.write(json.dumps(data, default=str).encode())


def _parse_date(s):
    """Parse YYYY-MM-DD string to datetime.date."""
    return datetime.datetime.strptime(s, '%Y-%m-%d').date()


def _safe_hours(stats, key='hours_worked'):
    """Extract a numeric hours value from a stats dict, defaulting to 0."""
    if not stats or not isinstance(stats, dict):
        return 0.0
    val = stats.get(key)
    if val is None or val == '':
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


# ── Pull logic ──────────────────────────────────────────────────────────────

def _pull_from_entries(user_id, pp_start):
    """
    Read det_entries for the 14-day pay period, split hours into reg/ot,
    auto-fill HL on holidays, upsert into timesheet_entries, return the rows.
    """
    sb = get_supabase()
    pp_start_date = _parse_date(pp_start)
    pp_end_date = pp_start_date + datetime.timedelta(days=13)
    holidays = _all_holidays()

    # Fetch det_entries for the 14 days
    resp = sb.table('det_entries') \
        .select('entry_date, stats') \
        .eq('user_id', user_id) \
        .gte('entry_date', pp_start) \
        .lte('entry_date', pp_end_date.isoformat()) \
        .execute()

    entry_map = {}
    for e in (resp.data or []):
        entry_map[e['entry_date']] = e.get('stats')

    rows = []
    for i in range(14):
        day = pp_start_date + datetime.timedelta(days=i)
        day_str = day.isoformat()
        stats = entry_map.get(day_str)
        total_hours = _safe_hours(stats)

        reg = total_hours

        code_rows = []
        if day in holidays:
            code_rows.append({'code': 'HL', 'hours': 8.0})

        rows.append({
            'user_id': user_id,
            'pay_period_start': pp_start,
            'day_date': day_str,
            'reg_hours': reg,
            'ot_hours': 0,
            'code_rows': code_rows,
            'auto_populated': True,
        })

    # Upsert all 14 rows
    sb.table('timesheet_entries') \
        .upsert(rows, on_conflict='user_id,pay_period_start,day_date') \
        .execute()

    return rows


def action_pull(body):
    """Pull: return existing rows or auto-populate from det_entries."""
    sb = get_supabase()
    user_id = body['user_id']
    pp_start = body['pay_period_start']

    # Check for existing rows
    resp = sb.table('timesheet_entries') \
        .select('*') \
        .eq('user_id', user_id) \
        .eq('pay_period_start', pp_start) \
        .order('day_date') \
        .execute()

    if resp.data and len(resp.data) == 14:
        return resp.data

    # No existing rows — pull fresh
    return _pull_from_entries(user_id, pp_start)


def action_save(body):
    """Save: upsert 14 day objects from the frontend."""
    sb = get_supabase()
    user_id = body['user_id']
    pp_start = body['pay_period_start']
    days = body['days']

    if len(days) != 14:
        raise ValueError(f'Expected 14 day objects, got {len(days)}')

    rows = []
    for d in days:
        rows.append({
            'user_id': user_id,
            'pay_period_start': pp_start,
            'day_date': d['day_date'],
            'reg_hours': float(d.get('reg_hours', 0)),
            'ot_hours': 0,
            'code_rows': d.get('code_rows', []),
            'auto_populated': False,
        })

    sb.table('timesheet_entries') \
        .upsert(rows, on_conflict='user_id,pay_period_start,day_date') \
        .execute()

    return rows


def action_reset(body):
    """Reset: delete existing rows, re-pull from det_entries."""
    sb = get_supabase()
    user_id = body['user_id']
    pp_start = body['pay_period_start']

    sb.table('timesheet_entries') \
        .delete() \
        .eq('user_id', user_id) \
        .eq('pay_period_start', pp_start) \
        .execute()

    return _pull_from_entries(user_id, pp_start)


def action_export(body):
    """Export: fill MASTER_BLANK.xlsx first detective block, return .xlsx bytes."""
    from openpyxl import load_workbook

    user_id = body['user_id']
    pp_start = body['pay_period_start']
    detective_name = body['detective_name']

    if not os.path.exists(MASTER_BLANK):
        raise FileNotFoundError(f'Template not found: {MASTER_BLANK}')

    # Fetch saved timesheet rows
    sb = get_supabase()
    resp = sb.table('timesheet_entries') \
        .select('*') \
        .eq('user_id', user_id) \
        .eq('pay_period_start', pp_start) \
        .order('day_date') \
        .execute()

    ts_rows = resp.data or []
    if len(ts_rows) != 14:
        raise ValueError(
            f'Expected 14 timesheet rows, got {len(ts_rows)}. '
            'Pull or save the timesheet before exporting.'
        )

    wb = load_workbook(MASTER_BLANK)
    ws = wb['JCCH1']

    # Date headers (JCCH2-5 reference JCCH1 via formulas)
    pp_start_date = _parse_date(pp_start)
    pp_end_date = pp_start_date + datetime.timedelta(days=13)
    ws['M1'] = pp_start_date
    ws['O1'] = pp_end_date

    # Fill first block using helper
    blocks = _find_blocks(ws)
    rg_row, code_rows_list, _total_row = blocks[0]
    _fill_detective_block(ws, rg_row, code_rows_list, ts_rows, detective_name)

    # Clear default placeholder 8s on unused blocks
    for _, (unused_rg, _unused_codes, _unused_total) in enumerate(blocks[1:]):
        _clear_block_defaults(ws, unused_rg)

    # Clear unused blocks on all other sheets
    for sheet_name in wb.sheetnames:
        if sheet_name == 'JCCH1':
            continue
        other_ws = wb[sheet_name]
        for unused_rg, _unused_codes, _unused_total in _find_blocks(other_ws):
            _clear_block_defaults(other_ws, unused_rg)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    last_name = detective_name.split()[-1]
    return output, last_name, pp_start


def action_export_all(body):
    """Export all detectives into MASTER_BLANK.xlsx, return .xlsx bytes."""
    from openpyxl import load_workbook

    pp_start = body['pay_period_start']

    if not os.path.exists(MASTER_BLANK):
        raise FileNotFoundError(f'Template not found: {MASTER_BLANK}')

    sb = get_supabase()

    # Fetch all users, sort by last name (case-insensitive)
    resp = sb.table('det_users').select('id, name').execute()
    users = sorted(
        resp.data or [],
        key=lambda u: u['name'].split()[-1].lower(),
    )

    wb = load_workbook(MASTER_BLANK)

    # Build flat list of (sheet, block) slots across all 5 sheets × 4 blocks
    all_slots = []  # [(ws, rg_row, code_rows, total_row), ...]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for rg_row, code_rows_list, total_row in _find_blocks(ws):
            all_slots.append((ws, rg_row, code_rows_list, total_row))

    max_slots = len(all_slots)
    if len(users) > max_slots:
        raise ValueError(
            f'Too many detectives ({len(users)}) for {max_slots} available '
            f'template slots. Maximum supported is {max_slots}.'
        )

    # Date headers on JCCH1 (other sheets reference via formulas)
    pp_start_date = _parse_date(pp_start)
    pp_end_date = pp_start_date + datetime.timedelta(days=13)
    wb['JCCH1']['M1'] = pp_start_date
    wb['JCCH1']['O1'] = pp_end_date

    # Fill each detective into their slot
    used_slots = set()
    for idx, user in enumerate(users):
        ws, rg_row, code_rows_list, _total_row = all_slots[idx]
        used_slots.add(idx)

        # Ensure timesheet data exists (pull from det_entries if needed)
        ts_resp = sb.table('timesheet_entries') \
            .select('*') \
            .eq('user_id', user['id']) \
            .eq('pay_period_start', pp_start) \
            .order('day_date') \
            .execute()

        if ts_resp.data and len(ts_resp.data) == 14:
            ts_rows = ts_resp.data
        else:
            ts_rows = _pull_from_entries(user['id'], pp_start)

        _fill_detective_block(ws, rg_row, code_rows_list, ts_rows, user['name'])

    # Clear default placeholder 8s from unused blocks
    for idx in range(len(all_slots)):
        if idx not in used_slots:
            ws, rg_row, _code_rows, _total_row = all_slots[idx]
            _clear_block_defaults(ws, rg_row)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, pp_start


# ── Handler ─────────────────────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = json.loads(self.rfile.read(content_length))

            action = body.get('action')
            if not action:
                _json_response(self, 400, {'error': 'Missing "action" field'})
                return

            if action == 'pull':
                result = action_pull(body)
                _json_response(self, 200, result)

            elif action == 'save':
                result = action_save(body)
                _json_response(self, 200, result)

            elif action == 'reset':
                result = action_reset(body)
                _json_response(self, 200, result)

            elif action == 'export':
                output, last_name, pp_start = action_export(body)
                filename = f'{last_name}_Timesheet_{pp_start}.xlsx'

                self.send_response(200)
                self.send_header(
                    'Content-Type',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                self.send_header(
                    'Content-Disposition',
                    f'attachment; filename="{filename}"'
                )
                _cors_headers(self)
                self.end_headers()
                self.wfile.write(output.read())

            elif action == 'export_all':
                output, pp_start = action_export_all(body)
                filename = f'JCSO_Detectives_Payroll_{pp_start}.xlsx'

                self.send_response(200)
                self.send_header(
                    'Content-Type',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                self.send_header(
                    'Content-Disposition',
                    f'attachment; filename="{filename}"'
                )
                _cors_headers(self)
                self.end_headers()
                self.wfile.write(output.read())

            else:
                _json_response(self, 400, {'error': f'Unknown action: {action}'})

        except Exception as exc:
            tb = traceback.format_exc()
            print(f'[timesheet] ERROR: {tb}')
            _json_response(self, 500, {'error': str(exc), 'traceback': tb})

    def do_OPTIONS(self):
        self.send_response(200)
        _cors_headers(self)
        self.end_headers()
