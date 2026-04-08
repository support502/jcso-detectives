"""
Vercel serverless function: /api/weekly
Accepts POST JSON with detective info + daily entries,
fills the correct Excel template, returns the completed .xlsx file.
"""
from http.server import BaseHTTPRequestHandler
import json, os, io, datetime
from openpyxl import load_workbook

# ── Paths to template files ──────────────────────────────────────────────────
TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), '..', 'templates')
TEMPLATES = {
    'Uniform':      os.path.join(TEMPLATE_DIR, '2026 Uniform Weekly.xlsx'),
    'UC':           os.path.join(TEMPLATE_DIR, 'Blank Narc Weekly.xlsx'),
    'Interdiction': os.path.join(TEMPLATE_DIR, 'Interdiction Weekly.xlsx'),
}

# ── Row numbers in both weekly templates (identical layout) ──────────────────
# Index 0 = Sunday, 6 = Saturday
DATA_ROWS    = [4, 7, 10, 13, 16, 19, 22]   # day data (stats in C–M)
ACTIVITY_ROWS = [5, 8, 11, 14, 17, 20, 23]   # "Activity:" notes in C (merged)

# ── Stats-key → column mapping per unit ──────────────────────────────────────
# Column order matches the template headers (C through M)
UNIFORM_COLS = [
    'hours_worked', 'time_off', 'k9_deploy', 'training_hours',
    'surv_hours', 'tno_cases', 'patrol_jail_cases', 'traffic_stops',
    'warrant_arrests', 'agency_assist', 'supp_reports',
]
UC_COLS = [
    'hours_worked', 'attempted_operations', 'uc_ci_cases', 'tno_cases',
    'sw_cases', 'surv_hours', 'patrol_jail_cases', 'pc_arrests',
    'warrant_arrests', 'training_hours', 'detective_agency_assist',
]
INTERDICTION_COLS = [
    'hours_worked', 'drug_seizures', 'criminal_seizures', 'currency_seizures',
    'training_hours', 'vehicle_searches', 'assist_narc_ops', 'traffic_stops',
    'warrant_arrests', 'pc_arrests', 'agency_assist',
]
UNIT_COLS = {
    'Uniform': UNIFORM_COLS,
    'UC': UC_COLS,
    'Interdiction': INTERDICTION_COLS,
}


def fill_template(unit, detective_name, week_start, entries):
    """
    Fill data into the weekly template for Uniform or UC.
    Returns a BytesIO object containing the completed .xlsx file.
    """
    template_path = TEMPLATES[unit]
    wb = load_workbook(template_path)
    ws = wb.active

    cols = UNIT_COLS[unit]
    # col_letters: C, D, E, ... for each stat
    col_letters = [chr(ord('C') + i) for i in range(len(cols))]

    # ── Write detective last name into A1 (no label prefix) ──
    last_name = detective_name.split()[-1]
    ws['A1'] = last_name

    # ── Write the actual date into every day's B cell ──
    # The template uses =B4+N formulas for Mon–Sat; openpyxl saves those as
    # literal strings that won't recalculate reliably. Replace all with real dates.
    sunday = datetime.datetime.strptime(week_start, '%Y-%m-%d')
    for day_idx, data_row in enumerate(DATA_ROWS):
        day_date = sunday + datetime.timedelta(days=day_idx)
        ws[f'B{data_row}'] = day_date

    # ── Build lookup: day_index (0=Sun … 6=Sat) → entry ──
    entry_map = {}
    for e in entries:
        d = datetime.datetime.strptime(e['entry_date'], '%Y-%m-%d')
        entry_map[d.weekday()] = e  # Python weekday: 0=Mon … 6=Sun

    # Convert Python weekday → our index (0=Sun)
    def py_to_our(py_wd):
        return (py_wd + 1) % 7  # Sun=6→0, Mon=0→1, … Sat=5→6

    day_entries = {}
    for py_wd, e in entry_map.items():
        day_entries[py_to_our(py_wd)] = e

    # ── Fill each day ──
    for day_idx in range(7):
        data_row = DATA_ROWS[day_idx]
        activity_row = ACTIVITY_ROWS[day_idx]
        entry = day_entries.get(day_idx)

        if entry:
            stats = entry.get('stats', {})
            # Fill stat cells (columns C onward)
            for i, key in enumerate(cols):
                cell = ws[f'{col_letters[i]}{data_row}']
                val = stats.get(key)
                cell.value = float(val) if val not in (None, '', 0, '0') else 0

            # Fill activity/notes row
            notes_parts = []
            if entry.get('case_numbers'):
                notes_parts.append(f"Cases: {entry['case_numbers']}")
            if entry.get('notes'):
                notes_parts.append(entry['notes'])
            if notes_parts:
                ws[f'C{activity_row}'] = '  |  '.join(notes_parts)
        else:
            # No entry for this day — write zeros for stat cells
            for i in range(len(cols)):
                cell = ws[f'{col_letters[i]}{data_row}']
                if cell.value is None or cell.value == 0:
                    cell.value = 0

    # ── Compute and write weekly totals to row 26 ──
    #
    # The template SUM formulas use two different inclusion patterns:
    #   cols[0] and cols[1]  (C and D)  →  Mon–Fri only  → rows 7,10,13,16,19
    #   cols[2] onward       (E–M)      →  all 7 days    → rows 4,7,10,13,16,19,22
    #
    # We calculate these sums from the values we just wrote, then overwrite the
    # formula cells with the numeric results. This is more reliable than relying
    # on Excel to recalculate the stale cached values that openpyxl carries over
    # from the original template.
    #
    # Verified against template formulas (both Uniform and UC are identical):
    #   C26 = '=SUM(C7,C10,C13,C16,C19)'
    #   D26 = '=SUM(D7,D10,D13,D16,D19)'
    #   E26 = '=SUM(E4,E7,E10,E13,E16,E19,E22)'  ← same pattern for F–M
    mon_fri_rows  = [7, 10, 13, 16, 19]           # day_idx 1–5
    all_seven_rows = [4, 7, 10, 13, 16, 19, 22]   # day_idx 0–6

    for i, cl in enumerate(col_letters):
        rows_to_sum = mon_fri_rows if i < 2 else all_seven_rows
        total = sum(ws[f'{cl}{r}'].value or 0 for r in rows_to_sum)
        ws[f'{cl}26'] = total

    # ── Save to bytes ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output



class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = json.loads(self.rfile.read(content_length))

            unit = body['unit']
            detective_name = body['detective_name']
            week_start = body['week_start']
            entries = body.get('entries', [])

            if unit in TEMPLATES:
                output = fill_template(unit, detective_name, week_start, entries)
            else:
                self.send_response(400)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': f'Unknown unit: {unit}'}).encode())
                return

            last_name = detective_name.split()[-1]
            filename = f'{last_name}Week{week_start.replace("-", "")}.xlsx'

            self.send_response(200)
            self.send_header('Content-Type',
                             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(output.read())

        except Exception as exc:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(exc)}).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
