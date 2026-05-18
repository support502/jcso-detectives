"""
Vercel serverless function: /api/monthly
Accepts POST JSON with month, year, and all entries for that month.
Opens the monthly Excel template, fills the correct month sheet, returns .xlsx.
"""
from http.server import BaseHTTPRequestHandler
import json, os, io, datetime, calendar, traceback
from openpyxl import load_workbook


# ── Supabase client (lazy-init) ─────────────────────────────────────────────
_supabase = None


def get_supabase():
    global _supabase
    if _supabase is None:
        from supabase import create_client
        url = os.environ.get('SUPABASE_URL')
        key = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')
        if not url or not key:
            raise RuntimeError(
                'Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY env vars. '
                'Set them in Vercel project settings.'
            )
        _supabase = create_client(url, key)
    return _supabase


def _compute_target_month(selected_month, year):
    """Upper bound for the month-loop in a monthly export.

    - Past year   → 12 (full year, the data is settled)
    - Current year → max(selected, today.month) (whichever is later)
    - Future year → just the selected month
    """
    today = datetime.date.today()
    if year < today.year:
        return 12
    if year > today.year:
        return selected_month
    return max(selected_month, today.month)


def _fetch_month_entries(year, month):
    """Pull det_entries for a single month. ~286 rows max, well under the
    Supabase default page size, so no pagination needed."""
    sb = get_supabase()
    last_day = calendar.monthrange(year, month)[1]
    start = datetime.date(year, month, 1).isoformat()
    end = datetime.date(year, month, last_day).isoformat()
    resp = sb.table('det_entries').select('*') \
        .gte('entry_date', start) \
        .lte('entry_date', end) \
        .execute()
    return resp.data or []

UNIFORM_TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), '..', 'templates', 'Monthly Uniform 2025 new.xlsx'
)
UC_TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), '..', 'templates', 'Monthly UC 2025 new.xlsx'
)

# ── Sheet names in the Uniform template (index 0–11 = Jan–Dec, 12 = Year Total) ──
SHEET_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# ── Sheet names in the UC template ──
UC_SHEET_NAMES = ['Jan', 'Feb', 'March', 'April', 'May', 'June',
                  'July', 'Aug', 'Sept', 'Oct', 'Nov', 'Dec']

# ═══════════════════════════════════════════════════════════════════════════
# CELL MAPPINGS — derived from studying the template
# Each detective → (name_row, week1_row … week5_row)
# The template already has SUM formulas in the total rows — we never touch those.
# ═══════════════════════════════════════════════════════════════════════════

# ── INTERDICTION section (rows 10–24) ──
# Stats columns B–L:
INTER_STAT_COLS = {
    'B': 'hours_worked',
    'C': 'drug_seizures',
    'D': 'criminal_seizures',
    'E': 'currency_seizures',
    'F': 'training_hours',
    'G': 'vehicle_searches',
    'H': 'assist_narc_ops',
    'I': 'traffic_stops',
    'J': 'warrant_arrests',
    'K': 'pc_arrests',
    'L': 'agency_assist',
}
# Drug seizure columns N–T:
INTER_DRUG_COLS = {
    'N': 'meth_g',
    'O': 'cocaine_g',
    'P': 'heroin_g',
    'Q': 'fentanyl_g',
    'R': 'marijuana_oz',
    'S': 'promethazine_codeine_oz',
    'T': 'rx_pills',
}

# Detective name → week rows + totals row for stats and drugs
INTER_DETECTIVES = {
    'Jake Droddy':    {'stat_weeks': [11, 12, 13, 14, 15], 'drug_weeks': [11, 12, 13, 14, 15],
                       'stat_total_row': 16, 'drug_total_row': 16},
    'Brigitte Morse': {'stat_weeks': [18, 19, 20, 21, 22], 'drug_weeks': [18, 19, 20, 21, 22],
                       'stat_total_row': 23, 'drug_total_row': 23},
}

# ── UNIFORM section (rows 27–55) ──
# Stats columns B–L:
UNI_STAT_COLS = {
    'B': 'hours_worked',
    'C': 'time_off',
    'D': 'k9_deploy',
    'E': 'directed_cases',
    'F': 'training_hours',
    'G': 'surv_hours',
    'H': 'self_made_cases',
    'I': 'traffic_stops',
    'J': 'warrant_arrests',
    'K': 'vehicle_searches',
    'L': 'supp_reports',
}

# Drug seizure columns N–T for Uniform detectives (right side of sheet):
# These use different rows than the stat columns
UNI_DETECTIVES = {
    'Scott Weaver':  {'stat_weeks': [28, 29, 30, 31, 32], 'drug_weeks': [25, 26, 27, 28, 29],
                      'stat_total_row': 33, 'drug_total_row': 30},
    'Brian Chowns':  {'stat_weeks': [35, 36, 37, 38, 39], 'drug_weeks': [32, 33, 34, 35, 36],
                      'stat_total_row': 40, 'drug_total_row': 37},
    'William Crain': {'stat_weeks': [42, 43, 44, 45, 46], 'drug_weeks': [39, 40, 41, 42, 43],
                      'stat_total_row': 47, 'drug_total_row': 44},
    'Tamara Spikes': {'stat_weeks': [49, 50, 51, 52, 53], 'drug_weeks': [46, 47, 48, 49, 50],
                      'stat_total_row': 54, 'drug_total_row': 51},
}

# ═══════════════════════════════════════════════════════════════════════════
# UC SECTION — separate template: Monthly UC 2025 new.xlsx
# ═══════════════════════════════════════════════════════════════════════════

# Left side activity stats — columns B through L
UC_STAT_COLS = {
    'B': 'hours_worked',
    'C': 'attempted_operations',
    'D': 'uc_ci_cases',
    'E': 'tno_cases',
    'F': 'sw_cases',
    'G': 'surv_hours',
    'H': 'patrol_jail_cases',
    'I': 'pc_arrests',
    'J': 'warrant_arrests',
    'K': 'training_hours',
    'L': 'detective_agency_assist',
}

# Right side drug seizures — columns N through T (detective name in column M)
UC_DRUG_COLS = {
    'N': 'meth_g',
    'O': 'cocaine_g',
    'P': 'heroin_g',
    'Q': 'fentanyl_g',
    'R': 'marijuana_oz',
    'S': 'promethazine_codeine_oz',
    'T': 'rx_pills',
}

# Detective name → start_row.  Week rows = start_row+1 … start_row+5, Total = start_row+6.
UC_DETECTIVES = {
    'Colton Lowe':    {'start_row': 3},
    'Layne Verdine':  {'start_row': 10},
    'Ryan Golmon':    {'start_row': 17},
    'Matthew Flowers': {'start_row': 24},
    'Tyler Lewis':     {'start_row': 31},
}


def get_week_start(date_str):
    """Return the Sunday that starts the week containing date_str."""
    d = datetime.date.fromisoformat(date_str)
    return d - datetime.timedelta(days=d.weekday() + 1 if d.weekday() != 6 else 0)


def get_week_starts_for_month(month, year):
    """
    Return up to 5 week-start Sundays that overlap with the given month.
    Ordered chronologically — index 0 = Week 1.
    """
    first = datetime.date(year, month, 1)
    # Sunday on or before the 1st
    start = first - datetime.timedelta(days=(first.weekday() + 1) % 7)
    last_day = (datetime.date(year, month % 12 + 1, 1) if month < 12
                else datetime.date(year + 1, 1, 1)) - datetime.timedelta(days=1)
    weeks = []
    current = start
    while len(weeks) < 6:
        week_end = current + datetime.timedelta(days=6)
        # Include if any day of this week falls in the target month
        if (current.month == month and current.year == year) or \
           (week_end.month == month and week_end.year == year):
            weeks.append(current.isoformat())
        current += datetime.timedelta(days=7)
        if current > last_day + datetime.timedelta(days=7):
            break
    return weeks[:5]


# ═══════════════════════════════════════════════════════════════════════════
# Year Total row mappings (Uniform template)
# ═══════════════════════════════════════════════════════════════════════════
INTER_YT = {
    'drug_seizures': 3, 'criminal_seizures': 4, 'currency_seizures': 5,
    'training_hours': 6, 'vehicle_searches': 7, 'assist_narc_ops': 8,
    'traffic_stops': 9, 'warrant_arrests': 10, 'pc_arrests': 11, 'agency_assist': 12,
}
UNI_YT = {
    'k9_deploy': 21, 'directed_cases': 22, 'training_hours': 23, 'surv_hours': 24,
    'self_made_cases': 25, 'traffic_stops': 26, 'warrant_arrests': 27,
    'vehicle_searches': 28, 'supp_reports': 29,
}
DRUG_YT = {
    'meth_g': 32, 'cocaine_g': 33, 'heroin_g': 34, 'fentanyl_g': 35,
    'marijuana_oz': 36, 'promethazine_codeine_oz': 37, 'rx_pills': 38,
}

# ═══════════════════════════════════════════════════════════════════════════
# Year Total row mappings (UC template)
# ═══════════════════════════════════════════════════════════════════════════
UC_YT_STATS = {
    'attempted_operations': 3, 'uc_ci_cases': 4, 'tno_cases': 5,
    'sw_cases': 6, 'surv_hours': 7, 'patrol_jail_cases': 8,
    'pc_arrests': 9, 'warrant_arrests': 10, 'training_hours': 11,
    'detective_agency_assist': 12,
}
UC_YT_DRUGS = {
    'meth_g': 17, 'cocaine_g': 18, 'heroin_g': 19, 'fentanyl_g': 20,
    'marijuana_oz': 21, 'promethazine_codeine_oz': 22, 'rx_pills': 23,
}


# ═══════════════════════════════════════════════════════════════════════════
# Shared helpers
# ═══════════════════════════════════════════════════════════════════════════

def _sum_stats(entry_list, key):
    """Sum a single stat key across a list of entries."""
    total = 0
    for e in entry_list:
        val = e.get('stats', {}).get(key)
        if val not in (None, '', '0'):
            try:
                total += float(val)
            except (ValueError, TypeError):
                pass
    return total if total else None


def _add(acc, key, val):
    if val:
        acc[key] = acc.get(key, 0.0) + val


def _build_week_entries(entries):
    """Build lookup: (user_name, week_start_str) → list of entries."""
    week_entries = {}
    for e in entries:
        ws_key = e.get('week_start', get_week_start(e['entry_date']).isoformat())
        key = (e['user_name'], ws_key)
        week_entries.setdefault(key, []).append(e)
    return week_entries


# ═══════════════════════════════════════════════════════════════════════════
# Uniform/Interdiction — fill a single month sheet on a workbook
# Returns (inter_month_acc, uni_month_acc, drug_month_acc)
# ═══════════════════════════════════════════════════════════════════════════

def _fill_uniform_month_sheet(wb, month, year, entries):
    """Fill one month sheet in the Uniform template. Returns accumulators."""
    sheet_name = SHEET_NAMES[month - 1]
    ws = wb[sheet_name]
    week_starts = get_week_starts_for_month(month, year)
    week_entries = _build_week_entries(entries)

    inter_month = {}
    uni_month = {}
    drug_month = {}

    def fill_detective(det_name, det_config, stat_cols, drug_cols=None,
                       stat_acc=None, drug_acc=None):
        det_stat = {}  # col_letter -> running total
        det_drug = {}
        for wi, ws_str in enumerate(week_starts):
            if wi >= 5:
                break
            elist = week_entries.get((det_name, ws_str), [])
            stat_row = det_config['stat_weeks'][wi]
            for col_letter, stat_key in stat_cols.items():
                val = _sum_stats(elist, stat_key)
                if val is not None:
                    ws[f'{col_letter}{stat_row}'] = val
                    det_stat[col_letter] = det_stat.get(col_letter, 0) + val
                    if stat_acc is not None:
                        _add(stat_acc, stat_key, val)
            if drug_cols:
                drug_row = det_config['drug_weeks'][wi]
                for col_letter, stat_key in drug_cols.items():
                    val = _sum_stats(elist, stat_key)
                    if val is not None:
                        ws[f'{col_letter}{drug_row}'] = val
                        det_drug[col_letter] = det_drug.get(col_letter, 0) + val
                        if drug_acc is not None:
                            _add(drug_acc, stat_key, val)
        # Write per-detective totals as static numbers (replacing SUM formulas)
        for col_letter, total in det_stat.items():
            ws[f'{col_letter}{det_config["stat_total_row"]}'] = total
        if drug_cols:
            for col_letter, total in det_drug.items():
                ws[f'{col_letter}{det_config["drug_total_row"]}'] = total
        return det_stat, det_drug

    # Interdiction — accumulate for unit-total row 24
    inter_stat_unit = {}
    inter_drug_all = {}
    for det_name, config in INTER_DETECTIVES.items():
        s, d = fill_detective(det_name, config, INTER_STAT_COLS, INTER_DRUG_COLS,
                              stat_acc=inter_month, drug_acc=drug_month)
        for col, v in s.items():
            inter_stat_unit[col] = inter_stat_unit.get(col, 0) + v
        for col, v in d.items():
            inter_drug_all[col] = inter_drug_all.get(col, 0) + v

    # Interdiction Unit Total (row 24, stat cols B–L)
    for col_letter, total in inter_stat_unit.items():
        ws[f'{col_letter}24'] = total

    # Uniform — accumulate for unit-total rows 52 and 55
    uni_stat_unit = {}
    uni_drug_all = {}
    for det_name, config in UNI_DETECTIVES.items():
        s, d = fill_detective(det_name, config, UNI_STAT_COLS, INTER_DRUG_COLS,
                              stat_acc=uni_month, drug_acc=drug_month)
        for col, v in s.items():
            uni_stat_unit[col] = uni_stat_unit.get(col, 0) + v
        for col, v in d.items():
            uni_drug_all[col] = uni_drug_all.get(col, 0) + v

    # Uniform Total (row 55, stat cols B–L)
    for col_letter, total in uni_stat_unit.items():
        ws[f'{col_letter}55'] = total

    # Grand drug total (row 52, drug cols N–T) — all 6 detectives combined
    grand_drug = {}
    for col, v in inter_drug_all.items():
        grand_drug[col] = grand_drug.get(col, 0) + v
    for col, v in uni_drug_all.items():
        grand_drug[col] = grand_drug.get(col, 0) + v
    for col_letter, total in grand_drug.items():
        ws[f'{col_letter}52'] = total

    return inter_month, uni_month, drug_month


def _write_uniform_year_total(wb, month, inter_month, uni_month, drug_month):
    """Write one month column to the Year Total sheet (Uniform template)."""
    yt = wb['Year Total']
    mc = chr(ord('B') + month - 1)
    for stat_key, yt_row in INTER_YT.items():
        val = inter_month.get(stat_key, 0)
        if val:
            yt[f'{mc}{yt_row}'] = val
    for stat_key, yt_row in UNI_YT.items():
        val = uni_month.get(stat_key, 0)
        if val:
            yt[f'{mc}{yt_row}'] = val
    for stat_key, yt_row in DRUG_YT.items():
        val = drug_month.get(stat_key, 0)
        if val:
            yt[f'{mc}{yt_row}'] = val


def fill_monthly(selected_month, year):
    """Fill Jan..target month sheets + Year Total in the Uniform template.

    target = max(selected_month, today.month) for the current year, 12 for
    past years, selected_month for future years. Months with no data are
    skipped, so the workbook reflects exactly what's in Supabase up to that
    bound. Returns BytesIO."""
    wb = load_workbook(UNIFORM_TEMPLATE_PATH)
    target = _compute_target_month(selected_month, year)

    for m in range(1, target + 1):
        month_entries = _fetch_month_entries(year, m)
        if not month_entries:
            continue
        inter_m, uni_m, drug_m = _fill_uniform_month_sheet(
            wb, m, year, month_entries)
        _write_uniform_year_total(wb, m, inter_m, uni_m, drug_m)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def fill_yearly(year, entries_by_month):
    """
    Fill ALL 12 month sheets + full Year Total (Uniform template).
    entries_by_month: dict {month_int: [entries]}
    """
    wb = load_workbook(UNIFORM_TEMPLATE_PATH)
    for m in range(1, 13):
        month_entries = entries_by_month.get(m, [])
        if not month_entries:
            continue
        inter_m, uni_m, drug_m = _fill_uniform_month_sheet(
            wb, m, year, month_entries)
        _write_uniform_year_total(wb, m, inter_m, uni_m, drug_m)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ═══════════════════════════════════════════════════════════════════════════
# UC — fill a single month sheet on a workbook
# Returns (uc_stat_acc, uc_drug_acc)
# ═══════════════════════════════════════════════════════════════════════════

def _fill_uc_month_sheet(wb, month, year, entries):
    """Fill one month sheet in the UC template. Returns accumulators."""
    sheet_name = UC_SHEET_NAMES[month - 1]
    ws = wb[sheet_name]
    week_starts = get_week_starts_for_month(month, year)
    week_entries = _build_week_entries(entries)

    uc_stat_month = {}
    uc_drug_month = {}
    unit_stat = {}  # col_letter -> total across all UC detectives
    unit_drug = {}

    for det_name, det_config in UC_DETECTIVES.items():
        start = det_config['start_row']
        total_row = start + 6  # template: name+1..+5 = weeks, +6 = Total
        det_stat = {}
        det_drug = {}

        for wi, ws_str in enumerate(week_starts):
            if wi >= 5:
                break
            elist = week_entries.get((det_name, ws_str), [])
            data_row = start + 1 + wi

            for col_letter, stat_key in UC_STAT_COLS.items():
                val = _sum_stats(elist, stat_key)
                if val is not None:
                    ws[f'{col_letter}{data_row}'] = val
                    det_stat[col_letter] = det_stat.get(col_letter, 0) + val
                    _add(uc_stat_month, stat_key, val)

            for col_letter, stat_key in UC_DRUG_COLS.items():
                val = _sum_stats(elist, stat_key)
                if val is not None:
                    ws[f'{col_letter}{data_row}'] = val
                    det_drug[col_letter] = det_drug.get(col_letter, 0) + val
                    _add(uc_drug_month, stat_key, val)

        # Write per-detective totals as static numbers (replacing SUM formulas)
        for col_letter, total in det_stat.items():
            ws[f'{col_letter}{total_row}'] = total
        for col_letter, total in det_drug.items():
            ws[f'{col_letter}{total_row}'] = total

        for col, v in det_stat.items():
            unit_stat[col] = unit_stat.get(col, 0) + v
        for col, v in det_drug.items():
            unit_drug[col] = unit_drug.get(col, 0) + v

    # Write UC Unit Total (row 38, replacing SUM formula)
    for col_letter, total in unit_stat.items():
        ws[f'{col_letter}38'] = total
    for col_letter, total in unit_drug.items():
        ws[f'{col_letter}38'] = total

    return uc_stat_month, uc_drug_month


def _write_uc_year_total(wb, month, uc_stat_month, uc_drug_month):
    """Write one month column to the Year Total sheet (UC template)."""
    yt = wb['Year Total ']  # NOTE: trailing space
    mc = chr(ord('B') + month - 1)
    for stat_key, yt_row in UC_YT_STATS.items():
        val = uc_stat_month.get(stat_key, 0)
        if val:
            yt[f'{mc}{yt_row}'] = val
    for stat_key, yt_row in UC_YT_DRUGS.items():
        val = uc_drug_month.get(stat_key, 0)
        if val:
            yt[f'{mc}{yt_row}'] = val


def fill_monthly_uc(selected_month, year):
    """Fill Jan..target month sheets in the UC template. UC's Year Total
    uses cross-sheet formulas, so it auto-recalculates from the month sheets
    when Excel opens the file — we don't write to it directly."""
    wb = load_workbook(UC_TEMPLATE_PATH)
    target = _compute_target_month(selected_month, year)

    for m in range(1, target + 1):
        month_entries = _fetch_month_entries(year, m)
        if not month_entries:
            continue
        _fill_uc_month_sheet(wb, m, year, month_entries)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def fill_yearly_uc(year, entries_by_month):
    """
    Fill ALL 12 month sheets (UC template).
    Year Total has cross-sheet refs — don't touch it.
    entries_by_month: dict {month_int: [entries]}
    """
    wb = load_workbook(UC_TEMPLATE_PATH)
    for m in range(1, 13):
        month_entries = entries_by_month.get(m, [])
        if not month_entries:
            continue
        _fill_uc_month_sheet(wb, m, year, month_entries)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = json.loads(self.rfile.read(content_length))

            year = int(body['year'])
            unit = body.get('unit', 'Uniform')
            export_type = body.get('export_type', 'month')

            month_names = [
                '', 'January', 'February', 'March', 'April', 'May', 'June',
                'July', 'August', 'September', 'October', 'November', 'December',
            ]

            if export_type == 'year':
                # entries_by_month: dict {month_int: [entries]}
                entries_by_month = body.get('entries_by_month', {})
                # Convert string keys to int (JSON keys are strings)
                entries_by_month = {int(k): v for k, v in entries_by_month.items()}
                if unit == 'UC':
                    output = fill_yearly_uc(year, entries_by_month)
                else:
                    output = fill_yearly(year, entries_by_month)
                filename = f'JCSO_Yearly_{unit}_{year}.xlsx'
            else:
                month = int(body['month'])
                # `entries` may be present in the payload from older clients
                # — ignore it; we fetch from Supabase ourselves so the export
                # spans Jan..target_month, not just the selected month.
                if unit == 'UC':
                    output = fill_monthly_uc(month, year)
                else:
                    output = fill_monthly(month, year)
                filename = f'JCSO_Monthly_{unit}_{month_names[month]}_{year}.xlsx'

            self.send_response(200)
            self.send_header('Content-Type',
                             'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(output.read())

        except Exception as exc:
            tb = traceback.format_exc()
            print(f'[monthly] ERROR: {tb}')
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(exc), 'traceback': tb}).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
