"""
Vercel serverless function: /api/ops_plans
CRUD for JCSO operational plans.

Actions: list, get, create, update, delete, submit, approve
"""
from http.server import BaseHTTPRequestHandler
import json, os, io, re, datetime, traceback


# ── Supabase client (lazy-init) ─────────────────────────────────────────────
_supabase = None

def get_supabase():
    global _supabase
    if _supabase is None:
        from supabase import create_client
        url = os.environ.get('SUPABASE_URL')
        key = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')
        if not url or not key:
            raise RuntimeError(
                'Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY env vars. '
                'Set them in Vercel project settings.'
            )
        _supabase = create_client(url, key)
    return _supabase


# ── Editable column whitelist ───────────────────────────────────────────────
# Fields the client may set on create/update. Server-managed fields
# (id, created_by, created_at, updated_at, status, supervisor_*) are
# excluded — they're set explicitly by the relevant actions.
EDITABLE_FIELDS = {
    'case_number', 'deconfliction', 'case_agent', 'operation_type',
    'city_county', 'briefing_datetime', 'operation_datetime',
    'background_info', 'synopsis',
    'briefing_address', 'briefing_city_state', 'briefing_zip', 'briefing_other',
    'operation_address', 'operation_city_state', 'operation_zip', 'operation_other',
    'suspects', 'residents', 'ci_uc_vehicles', 'personnel',
    'uc_arrest_signal', 'uc_no_response', 'uc_full_response',
    'uc_audible', 'uc_visual',
    'comms_radios', 'comms_channels', 'comms_cell_phones', 'comms_other',
    'monitoring_callyo', 'monitoring_1021',
    'monitoring_active', 'monitoring_active_channel',
    'agent_ci_contacts',
    'arrest_tbd', 'arrest_anticipated', 'arrest_charge',
    'arrest_not_anticipated', 'arrest_other',
    'medical_name', 'medical_address', 'medical_city_state',
    'medical_zip', 'medical_phone',
    'captain', 'lt_sergeant', 'contact_numbers',
    'media_contact_1', 'media_contact_2',
}


def _filter_editable(data):
    """Strip any keys not in the editable whitelist."""
    return {k: v for k, v in (data or {}).items() if k in EDITABLE_FIELDS}


def _now_iso():
    return datetime.datetime.now(datetime.timezone.utc).isoformat()


# ── Action handlers ─────────────────────────────────────────────────────────

def action_list(body):
    """List all ops_plans, newest first, with creator name joined in."""
    sb = get_supabase()

    plans_resp = sb.table('ops_plans') \
        .select('*') \
        .order('created_at', desc=True) \
        .execute()

    plans = plans_resp.data or []
    if not plans:
        return []

    # Resolve creator names in one query
    creator_ids = list({p['created_by'] for p in plans if p.get('created_by')})
    name_map = {}
    if creator_ids:
        users_resp = sb.table('det_users') \
            .select('id, name') \
            .in_('id', creator_ids) \
            .execute()
        name_map = {u['id']: u['name'] for u in (users_resp.data or [])}

    for p in plans:
        p['created_by_name'] = name_map.get(p.get('created_by'))

    return plans


def action_get(body):
    """Fetch a single plan by id."""
    sb = get_supabase()
    plan_id = body.get('id')
    if not plan_id:
        raise ValueError('Missing "id" field')

    resp = sb.table('ops_plans') \
        .select('*') \
        .eq('id', plan_id) \
        .limit(1) \
        .execute()

    rows = resp.data or []
    if not rows:
        raise ValueError(f'Plan not found: {plan_id}')

    plan = rows[0]
    if plan.get('created_by'):
        u = sb.table('det_users') \
            .select('name') \
            .eq('id', plan['created_by']) \
            .limit(1) \
            .execute()
        if u.data:
            plan['created_by_name'] = u.data[0]['name']

    return plan


def action_create(body):
    """Insert a new draft plan owned by created_by."""
    sb = get_supabase()
    created_by = body.get('created_by')
    if not created_by:
        raise ValueError('Missing "created_by" field')

    payload = _filter_editable(body.get('data'))
    payload['created_by'] = created_by
    payload['status'] = 'draft'

    resp = sb.table('ops_plans').insert(payload).execute()
    rows = resp.data or []
    if not rows:
        raise RuntimeError('Insert returned no rows')
    return rows[0]


def action_update(body):
    """Update editable fields on an existing plan; bumps updated_at."""
    sb = get_supabase()
    plan_id = body.get('id')
    if not plan_id:
        raise ValueError('Missing "id" field')

    payload = _filter_editable(body.get('data'))
    if not payload:
        raise ValueError('No editable fields supplied in "data"')
    payload['updated_at'] = _now_iso()

    resp = sb.table('ops_plans') \
        .update(payload) \
        .eq('id', plan_id) \
        .execute()

    rows = resp.data or []
    if not rows:
        raise ValueError(f'Plan not found or update affected 0 rows: {plan_id}')
    return rows[0]


def action_delete(body):
    """Hard-delete a plan, but only if status='draft'."""
    sb = get_supabase()
    plan_id = body.get('id')
    if not plan_id:
        raise ValueError('Missing "id" field')

    # Verify status before deleting
    check = sb.table('ops_plans') \
        .select('id, status') \
        .eq('id', plan_id) \
        .limit(1) \
        .execute()

    rows = check.data or []
    if not rows:
        raise ValueError(f'Plan not found: {plan_id}')
    if rows[0]['status'] != 'draft':
        raise ValueError(
            f"Cannot delete plan {plan_id}: status is "
            f"'{rows[0]['status']}', only 'draft' may be deleted"
        )

    sb.table('ops_plans').delete().eq('id', plan_id).execute()
    return {'deleted': plan_id}


def action_submit(body):
    """Mark plan as submitted."""
    sb = get_supabase()
    plan_id = body.get('id')
    if not plan_id:
        raise ValueError('Missing "id" field')

    resp = sb.table('ops_plans') \
        .update({'status': 'submitted', 'updated_at': _now_iso()}) \
        .eq('id', plan_id) \
        .execute()

    rows = resp.data or []
    if not rows:
        raise ValueError(f'Plan not found: {plan_id}')
    return rows[0]


# ── Excel export ────────────────────────────────────────────────────────────

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), '..', 'templates')
OPS_TEMPLATE = os.path.join(TEMPLATE_DIR, 'ops_plan_new.xlsx')

# Synopsis primary cell wraps inside merged B34:K46 (~13 rows). If text exceeds
# this many chars, the remainder is written into the secondary block (A333).
SYNOPSIS_PRIMARY_LIMIT = 1800

# Suspect/resident block field offsets from the block's TOP LABEL row.
# Every data cell sits one row below its label, so all offsets are odd.
# (col, row_offset_from_label, field_key)
PERSON_FIELDS = [
    ('D',  1, 'name'),
    ('D',  3, 'dob'),
    ('G',  3, 'age'),
    ('H',  3, 'race'),
    ('I',  3, 'sex'),
    ('J',  3, 'height'),
    ('K',  3, 'weight'),
    ('D',  5, 'address'),
    ('D',  7, 'city_state'),
    ('J',  7, 'zip'),
    ('D',  9, 'dl_number'),
    ('J',  9, 'dl_state'),
    ('D', 11, 'employer'),
    ('G', 11, 'occupation'),
    ('J', 11, 'employer_city_state'),
    ('D', 13, 'criminal_history'),
    ('J', 13, 'cautions'),
]

# Block label-row anchors (data is filled at label_row + offsets above).
SUSPECT_BLOCKS = [60, 191, 207, 223]
RESIDENT_BLOCKS = [76, 238, 254]

# Personnel: row 99 / 286 are LABEL rows; data starts one row below.
PERSONNEL_TABLE_1 = list(range(100, 119))  # rows 100..118 (19 rows)
PERSONNEL_TABLE_2 = list(range(287, 331))  # rows 287..330 (44 rows)
PERSONNEL_ROWS = PERSONNEL_TABLE_1 + PERSONNEL_TABLE_2
PERSONNEL_COLS = [('A', 'name_contact'), ('E', 'assignment'),
                  ('F', 'agency'), ('H', 'vehicle'), ('J', 'secondary')]


def _w(ws, cell, value):
    """Write to a cell only if value is non-empty/None. Lets template defaults
    survive when our data is blank."""
    if value is None or value == '':
        return
    ws[cell] = value


def _x(value):
    """Truthy → 'X', else None (skipped by _w)."""
    return 'X' if value else None


def _split_synopsis(text):
    """Return (primary_chunk, overflow_chunk). Splits at last whitespace before
    the primary limit so a word isn't cut in half."""
    if not text:
        return '', ''
    if len(text) <= SYNOPSIS_PRIMARY_LIMIT:
        return text, ''
    cut = text.rfind(' ', 0, SYNOPSIS_PRIMARY_LIMIT)
    if cut < SYNOPSIS_PRIMARY_LIMIT - 200:
        cut = SYNOPSIS_PRIMARY_LIMIT
    return text[:cut].rstrip(), text[cut:].lstrip()


def _fmt_signed_date(iso_str):
    """timestamptz → MM-DD-YYYY for the supervisor signed_at line."""
    if not iso_str:
        return ''
    try:
        d = datetime.datetime.fromisoformat(str(iso_str).replace('Z', '+00:00'))
        return d.strftime('%m-%d-%Y')
    except Exception:
        return str(iso_str)


def _filename_for(plan):
    """OPS_PLAN_{case}_{YYYY-MM-DD}.xlsx — sanitize case for filesystem use."""
    case = plan.get('case_number') or 'NO_CASE'
    case = re.sub(r'[^A-Za-z0-9._-]+', '_', str(case)).strip('_') or 'NO_CASE'
    date_str = ''
    for src in (plan.get('operation_datetime'), plan.get('created_at')):
        if not src:
            continue
        try:
            d = datetime.datetime.fromisoformat(str(src).replace('Z', '+00:00'))
            date_str = d.strftime('%Y-%m-%d')
            break
        except Exception:
            continue
    if not date_str:
        date_str = datetime.date.today().isoformat()
    return f'OPS_PLAN_{case}_{date_str}.xlsx'


def _fill_person_block(ws, base_row, person):
    """Fill one suspect/resident block at the given top row."""
    p = person or {}
    for col, off, key in PERSON_FIELDS:
        _w(ws, f'{col}{base_row + off}', p.get(key))


def _fill_personnel(ws, personnel):
    """Fill personnel rows; overflow into the second table at row 286."""
    if not personnel:
        return
    for i, p in enumerate(personnel):
        if i >= len(PERSONNEL_ROWS):
            break  # silently drop beyond available rows
        row = PERSONNEL_ROWS[i]
        for col, key in PERSONNEL_COLS:
            _w(ws, f'{col}{row}', (p or {}).get(key))


def _fill_ci_uc_vehicles(ws, vehicles):
    """Two fixed slots on label rows 93 and 95. Layout per row:
    [B: CI checkbox] [C='CI' label] [D: UC checkbox] [E='UC' label]
    [F:H='Vehicle/LP:' label] [I:L data]."""
    rows = [93, 95]
    for i, row in enumerate(rows):
        if i >= len(vehicles or []):
            break
        v = vehicles[i] or {}
        t = (v.get('type') or '').upper()
        if t == 'CI':
            _w(ws, f'B{row}', 'X')
        elif t == 'UC':
            _w(ws, f'D{row}', 'X')
        _w(ws, f'I{row}', v.get('vehicle_lp'))


def _fill_agent_ci_contacts(ws, contacts):
    """Two slots — type rows 146/149 (with name), number rows 147/150.
    Layout per type row:
    [B: CI checkbox] [C='CI' label] [D: UC checkbox] [E='UC' label]
    [F='Name:' label] [H..L data].
    Number row: [F='Number:' label] [H..L data]."""
    slots = [(146, 147), (149, 150)]
    for i, (type_row, num_row) in enumerate(slots):
        if i >= len(contacts or []):
            break
        c = contacts[i] or {}
        t = (c.get('type') or '').upper()
        if t == 'CI':
            _w(ws, f'B{type_row}', 'X')
        elif t == 'UC':
            _w(ws, f'D{type_row}', 'X')
        _w(ws, f'H{type_row}', c.get('name'))
        _w(ws, f'H{num_row}', c.get('number'))


def _fill_media_contact(ws, data_row, m):
    """data_row is the row BELOW the label row. Name → A, Title → E, Phone → I."""
    if not m:
        return
    _w(ws, f'A{data_row}', m.get('name'))
    _w(ws, f'E{data_row}', m.get('title'))
    _w(ws, f'I{data_row}', m.get('phone'))


def _build_workbook(plan):
    """Open the template, write every field, return an in-memory buffer."""
    from openpyxl import load_workbook

    if not os.path.exists(OPS_TEMPLATE):
        raise FileNotFoundError(
            f'Ops plan template not found at {OPS_TEMPLATE}. '
            'Add templates/ops_plan_new.xlsx to the repo and redeploy.'
        )

    wb = load_workbook(OPS_TEMPLATE, data_only=False)
    ws = wb.active

    # ── Case info ────────────────────────────────────────────────────────
    # Row 11/12: horizontal layout — label in D:F merge, data in G:I merge.
    _w(ws, 'G11', plan.get('case_number'))
    _w(ws, 'G12', plan.get('deconfliction'))
    # Row 14: labels A/E/I; data row 15.
    _w(ws, 'A15', plan.get('case_agent'))
    _w(ws, 'E15', plan.get('operation_type'))
    _w(ws, 'I15', plan.get('city_county'))
    # Row 17: labels A/G; data row 18.
    _w(ws, 'A18', plan.get('briefing_datetime'))
    _w(ws, 'G18', plan.get('operation_datetime'))

    # ── Background / synopsis ────────────────────────────────────────────
    # Label rows B20 / B34; data rows start at B21 / B35 inside large merges.
    _w(ws, 'B21', plan.get('background_info'))
    primary, overflow = _split_synopsis(plan.get('synopsis') or '')
    _w(ws, 'B35', primary)
    if overflow:
        # Secondary synopsis area: label A332:L332, data A333:L379.
        _w(ws, 'A333', overflow)

    # ── Briefing location ────────────────────────────────────────────────
    # Row 48 labels Address/City/Zip; data row 49. Row 50 label "Other"; data 51.
    _w(ws, 'A49', plan.get('briefing_address'))
    _w(ws, 'H49', plan.get('briefing_city_state'))
    _w(ws, 'K49', plan.get('briefing_zip'))
    _w(ws, 'A51', plan.get('briefing_other'))

    # ── Operation location ───────────────────────────────────────────────
    _w(ws, 'A55', plan.get('operation_address'))
    _w(ws, 'H55', plan.get('operation_city_state'))
    _w(ws, 'K55', plan.get('operation_zip'))
    _w(ws, 'A57', plan.get('operation_other'))

    # ── Suspects (up to 4) ───────────────────────────────────────────────
    suspects = plan.get('suspects') or []
    for i, base in enumerate(SUSPECT_BLOCKS):
        if i >= len(suspects):
            break
        _fill_person_block(ws, base, suspects[i])

    # ── Residents (up to 3) ──────────────────────────────────────────────
    residents = plan.get('residents') or []
    for i, base in enumerate(RESIDENT_BLOCKS):
        if i >= len(residents):
            break
        _fill_person_block(ws, base, residents[i])

    # ── CI/UC vehicles ───────────────────────────────────────────────────
    _fill_ci_uc_vehicles(ws, plan.get('ci_uc_vehicles') or [])

    # ── Personnel (overflows into secondary table) ───────────────────────
    _fill_personnel(ws, plan.get('personnel') or [])

    # ── UC signals ───────────────────────────────────────────────────────
    # Row 121 has 3 column-header labels (Arrest Signal / No Response / Full
    # Response). Rows 123 and 126 are the row-label rows (Audible: / Visual:)
    # with merged data cells alongside. Schema has 5 standalone string fields,
    # so place each below or beside its closest label.
    _w(ws, 'A122', plan.get('uc_arrest_signal'))
    _w(ws, 'E122', plan.get('uc_no_response'))
    _w(ws, 'I122', plan.get('uc_full_response'))
    _w(ws, 'B123', plan.get('uc_audible'))   # right of "Audible:" label
    _w(ws, 'B126', plan.get('uc_visual'))    # right of "Visual:" label

    # ── Communications ───────────────────────────────────────────────────
    # Checkboxes sit in column B (left of their label in column C).
    _w(ws, 'B132', _x(plan.get('comms_radios')))
    _w(ws, 'G132', plan.get('comms_channels'))   # right of "Channel(s):" (E132:F132 label)
    _w(ws, 'B133', _x(plan.get('comms_cell_phones')))
    _w(ws, 'D134', plan.get('comms_other'))      # right of "Other:" (C134 label)

    # ── Monitoring ───────────────────────────────────────────────────────
    _w(ws, 'B139', _x(plan.get('monitoring_callyo')))
    _w(ws, 'B140', _x(plan.get('monitoring_1021')))
    _w(ws, 'B141', _x(plan.get('monitoring_active')))
    _w(ws, 'F141', plan.get('monitoring_active_channel'))  # right of "Active Channel:" label

    # ── Agent/CI contacts ────────────────────────────────────────────────
    _fill_agent_ci_contacts(ws, plan.get('agent_ci_contacts') or [])

    # ── Arrest ───────────────────────────────────────────────────────────
    _w(ws, 'B155', _x(plan.get('arrest_tbd')))
    _w(ws, 'B156', _x(plan.get('arrest_anticipated')))
    _w(ws, 'F156', plan.get('arrest_charge'))    # right of "Charge:" (E156 label)
    _w(ws, 'B157', _x(plan.get('arrest_not_anticipated')))
    _w(ws, 'D158', plan.get('arrest_other'))     # right of "Other:" (C158 label)

    # ── Medical ──────────────────────────────────────────────────────────
    # All medical labels are on rows 163/165; data rows are 164/166.
    _w(ws, 'A164', plan.get('medical_name'))
    _w(ws, 'G164', plan.get('medical_address'))
    _w(ws, 'A166', plan.get('medical_city_state'))
    _w(ws, 'E166', plan.get('medical_zip'))
    _w(ws, 'G166', plan.get('medical_phone'))

    # ── Division supervisors ─────────────────────────────────────────────
    _w(ws, 'A170', plan.get('captain'))
    _w(ws, 'G170', plan.get('lt_sergeant'))

    # ── Contact numbers ──────────────────────────────────────────────────
    # A171:L171 is the section label; A172:F172 is the data merge.
    _w(ws, 'A172', plan.get('contact_numbers'))

    # ── Media inquiries ──────────────────────────────────────────────────
    # Label rows 175 / 177 → data rows 176 / 178.
    _fill_media_contact(ws, 176, plan.get('media_contact_1') or {})
    _fill_media_contact(ws, 178, plan.get('media_contact_2') or {})

    # ── Supervisor approval ──────────────────────────────────────────────
    # Horizontal layout: A183:B183 label "Signature" → C183:K183 data;
    # A186:B186 label "Date" → C186 data; G186:H186 label "Rank" → I186 data.
    _w(ws, 'C183', plan.get('supervisor_signature'))
    _w(ws, 'C186', _fmt_signed_date(plan.get('supervisor_signed_at')))
    _w(ws, 'I186', plan.get('supervisor_rank'))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def action_export(body):
    """Build the filled .xlsx for a plan id. Returns (BytesIO, filename)."""
    sb = get_supabase()
    plan_id = body.get('id')
    if not plan_id:
        raise ValueError('Missing "id" field')

    resp = sb.table('ops_plans').select('*').eq('id', plan_id).limit(1).execute()
    rows = resp.data or []
    if not rows:
        raise ValueError(f'Plan not found: {plan_id}')
    plan = rows[0]

    output = _build_workbook(plan)
    return output, _filename_for(plan)


def action_approve(body):
    """Approve a plan: set supervisor signature, rank, signed_at, status."""
    sb = get_supabase()
    plan_id = body.get('id')
    if not plan_id:
        raise ValueError('Missing "id" field')

    signature = body.get('supervisor_signature')
    rank = body.get('supervisor_rank')
    if not signature:
        raise ValueError('Missing "supervisor_signature" field')
    if not rank:
        raise ValueError('Missing "supervisor_rank" field')

    now = _now_iso()
    resp = sb.table('ops_plans').update({
        'status': 'approved',
        'supervisor_signature': signature,
        'supervisor_signed_at': now,
        'supervisor_rank': rank,
        'updated_at': now,
    }).eq('id', plan_id).execute()

    rows = resp.data or []
    if not rows:
        raise ValueError(f'Plan not found: {plan_id}')
    return rows[0]


# ── HTTP helpers ────────────────────────────────────────────────────────────

def _cors_headers(h):
    h.send_header('Access-Control-Allow-Origin', '*')
    h.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
    h.send_header('Access-Control-Allow-Headers', 'Content-Type')


def _json_response(h, status, data):
    h.send_response(status)
    h.send_header('Content-Type', 'application/json')
    _cors_headers(h)
    h.end_headers()
    h.wfile.write(json.dumps(data, default=str).encode())


# ── Handler ─────────────────────────────────────────────────────────────────

ACTIONS = {
    'list': action_list,
    'get': action_get,
    'create': action_create,
    'update': action_update,
    'delete': action_delete,
    'submit': action_submit,
    'approve': action_approve,
}


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = json.loads(self.rfile.read(content_length)) if content_length else {}

            action = body.get('action')
            if not action:
                _json_response(self, 400, {'error': 'Missing "action" field'})
                return

            # Binary download path — handled separately from JSON actions
            if action == 'export':
                output, filename = action_export(body)
                self.send_response(200)
                self.send_header(
                    'Content-Type',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                self.send_header(
                    'Content-Disposition',
                    f'attachment; filename="{filename}"'
                )
                _cors_headers(self)
                self.end_headers()
                self.wfile.write(output.read())
                return

            fn = ACTIONS.get(action)
            if not fn:
                _json_response(self, 400, {'error': f'Unknown action: {action}'})
                return

            result = fn(body)
            _json_response(self, 200, result)

        except ValueError as exc:
            print(f'[ops_plans] BAD REQUEST: {exc}')
            _json_response(self, 400, {'error': str(exc)})

        except Exception as exc:
            tb = traceback.format_exc()
            print(f'[ops_plans] ERROR: {tb}')
            _json_response(self, 500, {'error': str(exc), 'traceback': tb})

    def do_OPTIONS(self):
        self.send_response(200)
        _cors_headers(self)
        self.end_headers()
